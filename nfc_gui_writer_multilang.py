
import nfc
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook
import threading

languages = {
    'zh': {
        'title': 'NFC 名片批次寫入工具',
        'select_file': '請選擇 Excel 檔案：',
        'choose': '選擇檔案',
        'start': '初始化 NFC 寫入',
        'write_next': '➡️ 寫入下一筆',
        'done': '✅ 寫入完成',
        'success': '✅ 寫入成功',
        'fail': '⚠️ 寫入失敗',
        'scan': '📇 請將 NFC 標籤靠近：',
        'error': '錯誤',
        'lang_switch': '切換語言'
    },
    'en': {
        'title': 'NFC vCard Batch Writer',
        'select_file': 'Select Excel file:',
        'choose': 'Browse',
        'start': 'Initialize NFC Writing',
        'write_next': '➡️ Write Next',
        'done': '✅ Writing Complete',
        'success': '✅ Write successful',
        'fail': '⚠️ Write failed',
        'scan': '📇 Hold NFC tag near reader for: ',
        'error': 'Error',
        'lang_switch': 'Switch Language'
    }
}

current_lang = 'zh'
rows = []
row_index = 0

def generate_vcard(name, title, phone, email, company, address, website):
    return f"""BEGIN:VCARD
VERSION:3.0
FN:{name}
ORG:{company}
TITLE:{title}
TEL:{phone}
EMAIL:{email}
ADR:{address}
URL:{website}
END:VCARD"""

def write_to_tag(vcard_text, log_text):
    try:
        clf = nfc.ContactlessFrontend('usb')
        def connected(tag):
            if tag.ndef:
                record = nfc.ndef.TextRecord(vcard_text)
                tag.ndef.records = [record]
                return True
            return False
        clf.connect(rdwr={'on-connect': connected})
        clf.close()
        return True
    except Exception as e:
        log_text.insert(tk.END, f"❌ {str(e)}\n")
        log_text.update()
        return False

def load_excel(file_path, log_text):
    global rows, row_index
    try:
        from openpyxl import load_workbook
        wb = load_workbook(filename=file_path)
        sheet = wb.active
        rows = list(sheet.iter_rows(min_row=2, values_only=True))
        row_index = 0
        log_text.insert(tk.END, f"📁 共載入 {len(rows)} 筆資料，請點選『寫入下一筆』開始寫入 NFC\n\n")
        log_text.update()
    except Exception as e:
        messagebox.showerror(languages[current_lang]['error'], str(e))

def write_next_entry(log_text):
    global row_index
    strings = languages[current_lang]
    if row_index >= len(rows):
        log_text.insert(tk.END, f"{strings['done']}\n")
        return

    row = rows[row_index]
    name, title, phone, email, company, address, website = row
    vcard = generate_vcard(name, title, phone, email, company, address, website)
    log_text.insert(tk.END, f"{strings['scan']}{name} (第 {row_index + 2} 行)...\n")
    log_text.update()
    result = write_to_tag(vcard, log_text)
    if result:
        log_text.insert(tk.END, f"{strings['success']}\n\n")
    else:
        log_text.insert(tk.END, f"{strings['fail']}\n\n")
    row_index += 1
    log_text.update()

def browse_file(entry, log_text):
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    entry.delete(0, tk.END)
    entry.insert(0, file_path)
    load_excel(file_path, log_text)

def run_gui():
    global current_lang
    window = tk.Tk()
    window.geometry("700x500")

    def refresh_ui():
        window.title(languages[current_lang]['title'])
        label.config(text=languages[current_lang]['select_file'])
        button_browse.config(text=languages[current_lang]['choose'])
        button_start.config(text=languages[current_lang]['start'])
        button_next.config(text=languages[current_lang]['write_next'])
        button_lang.config(text=languages[current_lang]['lang_switch'])

    def switch_language():
        global current_lang
        current_lang = 'en' if current_lang == 'zh' else 'zh'
        refresh_ui()

    label = tk.Label(window)
    label.pack(pady=5)

    entry = tk.Entry(window, width=60)
    entry.pack(pady=5)

    log_text = tk.Text(window, height=17)
    log_text.pack(pady=10)

    button_browse = tk.Button(window, command=lambda: browse_file(entry, log_text))
    button_browse.pack(pady=5)

    button_start = tk.Button(window, bg="blue", fg="white", command=lambda: load_excel(entry.get(), log_text))
    button_start.pack(pady=5)

    button_next = tk.Button(window, bg="green", fg="white", command=lambda: write_next_entry(log_text))
    button_next.pack(pady=5)

    button_lang = tk.Button(window, command=switch_language)
    button_lang.pack(pady=5)

    refresh_ui()
    window.mainloop()

if __name__ == "__main__":
    run_gui()
